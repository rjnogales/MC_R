"""
06_generar_excel.py
Lee los 4 CSV del temp y genera el Excel final con 4 hojas:
  - Inputs
  - Outputs
  - Variables_derivadas
  - Relaciones

Genera: <carpeta_salida>/<nombre_reporte>_diccionario.xlsx

Uso:
    python 06_generar_excel.py <carpeta_temp> <carpeta_salida> <nombre_reporte>
"""

import sys
import csv
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR 06] Falta openpyxl. Instalar con: pip install openpyxl")
    sys.exit(1)


COLORES_HOJA = {
    "Inputs":              "4472C4",   # azul
    "Outputs":             "ED7D31",   # naranja
    "Variables_derivadas": "70AD47",   # verde
    "Relaciones":          "9E480E",   # marrón
    "Glosario":            "7030A0",   # púrpura
}

SHEET_DESCRIPTIONS = {
    "Inputs": "Lista de controles definidas en la UI de Shiny.",
    "Outputs": "Salidas renderizadas por el servidor Shiny.",
    "Variables_derivadas": "Variables calculadas en preprocesamiento a partir de datasets principales.",
    "Relaciones": "Dependencias entre outputs y los inputs que los alimentan.",
    "Glosario": "Descripción de cada hoja y definición de las columnas incluidas en el diccionario.",
}

COLUMN_DESCRIPTIONS = {
    "Inputs": {
        "id": "Identificador único del control en el UI.",
        "tipo": "Tipo de control de Shiny (selectInput, checkboxInput, etc.).",
        "etiqueta": "Texto visible asociado al control en la interfaz.",
        "opciones": "Opciones disponibles para el input o tipo de valor esperado.",
    },
    "Outputs": {
        "id": "Identificador único del output en el servidor.",
        "tipo": "Tipo de render usado por el output (renderPlot, downloadHandler, etc.).",
    },
    "Variables_derivadas": {
        "dataset": "Nombre del objeto o dataset en el código R donde se crea la variable.",
        "columna_nueva": "Nombre de la columna o variable calculada.",
        "formula": "Expresión usada para calcular la variable derivada.",
        "linea": "Número de línea en el script donde se detectó la asignación.",
    },
    "Relaciones": {
        "output_id": "Identificador del output Shiny al que se refiere la relación.",
        "inputs_usados": "Lista de inputs usados dentro del bloque de renderizado de ese output.",
    },
}

INPUT_DESCRIPTIONS = {
    "Tipo_Evento": "Tipo de novedad o evento de la bitácora (accidente, vandalismo, incidente, etc.).",
    "COT1": "Concesionario u operador seleccionado para filtrar los datos.",
    "Tipo_Salida2": "Tipo de salida seleccionado para filtrar los registros.",
    "Seleccion": "Modo de selección para agrupar datos (COT, TIPOLOGIA, SEVERIDAD, etc.).",
    "COT3": "Concesionario u operador usado para el gráfico de jerarquías.",
    "Jerarquias3": "Jerarquía de ruta o servicio para el gráfico de jerarquías.",
    "Puntos1": "Activa el uso de valores de puntos en el gráfico.",
    "Cantidad1": "Activa el uso de conteos (cantidad de eventos) en el gráfico.",
    "Desvios1": "Incluye la serie de desviaciones en el gráfico de tiempos.",
    "Adicionales1": "Incluye la serie de datos adicionales en el gráfico de tiempos.",
    "N.A1": "Incluye valores N.A. en la visualización de serie de eventos.",
    "Incidente1": "Incluye la serie de incidentes en el gráfico de tiempos.",
    "Salida1": "Incluye la serie de salidas en el gráfico de tiempos.",
    "Ingreso1": "Incluye la serie de ingresos en el gráfico de tiempos.",
    "Inspección1": "Incluye la serie de inspecciones en el gráfico de tiempos.",
    "Afectaciones": "Agrupa los datos por ruta para el análisis de cantidad o puntos.",
    "Afectaciones1": "Agrupa los datos por número de vehículo para el análisis.",
    "Afectaciones3": "Agrupa los datos por código de operador para el análisis.",
    "Afectaciones4": "Agrupa los datos por quincena / persona que registra para el análisis.",
    "Afectaciones5": "Agrupa los datos por día para el análisis.",
    "Afectaciones6": "Agrupa los datos por hora para el análisis.",
    "Afectaciones_2": "Muestra los datos de accidentes en el gráfico o descarga.",
    "Afectaciones1_2": "Muestra los datos de vandalismo en el gráfico o descarga.",
    "Afectaciones2_2": "Muestra los puntos IE/IO en el gráfico o descarga.",
    "Afectaciones3_2": "Muestra los puntos IE en el gráfico o descarga.",
    "Afectaciones4_2": "Muestra métricas de Dane o Bitácora en el gráfico o descarga.",
    "Afectaciones5_2": "Muestra los reposicionamientos en el gráfico o descarga.",
    "Afectaciones6_2": "Muestra los incidentes en el gráfico o descarga.",
    "Afectaciones7_2": "Muestra los hitos en el gráfico o descarga.",
    "Afectaciones8_2": "Muestra kilómetros perdidos por salidas en el gráfico o descarga.",
    "Afectaciones9_2": "Muestra flota promedio por tableros en el gráfico o descarga.",
    "Afectaciones10_2": "Muestra flota promedio por tipología en el gráfico o descarga.",
    "Puntos": "Activa el uso de puntos agregados en selecciones COT/TIPOLOGIA.",
    "Cantidad": "Activa el uso de cantidades agregadas en selecciones COT/TIPOLOGIA.",
    "ART3": "Filtra la visualización por tipología ART.",
    "PAD3": "Filtra la visualización por tipología PAD.",
    "COM3": "Filtra la visualización por tipología COM.",
    "DUAL3": "Filtra la visualización por tipología DUAL.",
    "descarga": "Botón para descargar datos en formato CSV según los filtros activos.",
    "descarga2": "Botón para descargar una base de datos en formato Excel según los filtros activos.",
}

OUTPUT_DESCRIPTIONS = {
    "distPlot": "Gráfico de tipo de novedad por ruta/jerarquía en el rango de fechas seleccionado.",
    "distPlot2": "Gráfico de rutas con mayor cantidad de eventos para el tipo de novedad seleccionado.",
    "distPlot1": "Gráfico de la evolución de eventos por agrupación y series seleccionadas.",
    "distPlot1_2": "Gráfico de flota promedio o métricas agregadas según el filtro seleccionado.",
    "distPlot1_3": "Gráfico resumen por selección de COT, tipología o severidad.",
    "distPlot3": "Gráfico de incidencias por jerarquía y concesionario.",
    "descarga": "Descarga un CSV con los datos filtrados actualmente.",
    "descarga2": "Descarga un archivo Excel con la base de datos filtrada actualmente.",
}


def leer_csv(ruta: Path) -> tuple:
    """Devuelve (cabeceras, filas) de un CSV"""
    if not ruta.exists():
        return [], []
    with open(ruta, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        filas = list(reader)
        cabeceras = reader.fieldnames or []
    return cabeceras, filas


def escribir_hoja(ws, cabeceras: list, filas: list, color_hex: str):
    """Escribe cabeceras con formato y filas de datos"""
    fill_header = PatternFill("solid", fgColor=color_hex)
    font_header = Font(bold=True, color="FFFFFF")
    font_datos  = Font(name="Calibri", size=10)

    # Cabeceras
    for col_idx, nombre in enumerate(cabeceras, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, nombre_col in enumerate(cabeceras, start=1):
            celda = ws.cell(row=row_idx, column=col_idx, value=fila.get(nombre_col, ""))
            celda.font = font_datos
            celda.alignment = Alignment(wrap_text=True, vertical="top")

    # Ancho automático por columna
    for col_idx, nombre in enumerate(cabeceras, start=1):
        max_ancho = len(nombre)
        for fila in filas:
            valor = str(fila.get(nombre, ""))
            # Para celdas largas limitar el ancho visual
            max_ancho = max(max_ancho, min(len(valor), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_ancho + 4

    # Congelar primera fila
    ws.freeze_panes = "A2"


def escribir_glosario(wb, carpeta_temp: Path):
    ws = wb.create_sheet(title="Glosario")
    ws.sheet_view.tabSelected = False

    # Sección 1: descripción de las hojas
    ws.cell(row=1, column=1, value="Hoja")
    ws.cell(row=1, column=2, value="Descripción")
    for idx, (nombre_hoja, descripcion) in enumerate(SHEET_DESCRIPTIONS.items(), start=2):
        ws.cell(row=idx, column=1, value=nombre_hoja)
        ws.cell(row=idx, column=2, value=descripcion)

    inicio_columnas = len(SHEET_DESCRIPTIONS) + 4
    ws.cell(row=inicio_columnas - 1, column=1, value="Hoja")
    ws.cell(row=inicio_columnas - 1, column=2, value="Columna")
    ws.cell(row=inicio_columnas - 1, column=3, value="Descripción")

    fila = inicio_columnas
    for hoja, columnas in COLUMN_DESCRIPTIONS.items():
        for columna, descripcion in columnas.items():
            ws.cell(row=fila, column=1, value=hoja)
            ws.cell(row=fila, column=2, value=columna)
            ws.cell(row=fila, column=3, value=descripcion)
            fila += 1

    fila += 2
    # Sección 2: definiciones de elementos específicos
    ws.cell(row=fila, column=1, value="Elemento")
    ws.cell(row=fila, column=2, value="Tipo")
    ws.cell(row=fila, column=3, value="Descripción")
    fila += 1

    def escribir_elementos(cabeceras, filas, tipo):
        nonlocal fila
        for valor in filas:
            elemento = valor.get(cabeceras[0], "")
            descripcion = ""
            if tipo == "Inputs":
                descripcion = INPUT_DESCRIPTIONS.get(elemento, valor.get("etiqueta", ""))
            elif tipo == "Outputs":
                descripcion = OUTPUT_DESCRIPTIONS.get(elemento, "")
            if not descripcion:
                descripcion = "Descripción no disponible, revisar el script o el label asociado."
            ws.cell(row=fila, column=1, value=elemento)
            ws.cell(row=fila, column=2, value=tipo)
            ws.cell(row=fila, column=3, value=descripcion)
            fila += 1

    for hoja in ["Inputs", "Outputs"]:
        cabeceras, filas = leer_csv(carpeta_temp / f"{hoja.lower()}.csv")
        if cabeceras and filas:
            escribir_elementos(cabeceras, filas, hoja)

    for col_idx in range(1, 4):
        cell = ws.cell(row=inicio_columnas - 1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="444444")
        cell.alignment = Alignment(horizontal="center")

    # Ajustar ancho de las columnas del glosario
    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 36

    ws.freeze_panes = "A2"


def generar_excel(carpeta_temp: Path, carpeta_reporte: Path, nombre_reporte: str):
    hojas = {
        "Inputs":              carpeta_temp / "inputs.csv",
        "Outputs":             carpeta_temp / "outputs.csv",
        "Variables_derivadas": carpeta_temp / "variables_derivadas.csv",
        "Relaciones":          carpeta_temp / "relaciones.csv",
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto
    escribir_glosario(wb, carpeta_temp)

    for nombre_hoja, ruta_csv in hojas.items():
        cabeceras, filas = leer_csv(ruta_csv)
        if not cabeceras:
            print(f"        [AVISO] Sin datos para hoja: {nombre_hoja}")
            cabeceras, filas = [nombre_hoja], [{}]

        ws = wb.create_sheet(title=nombre_hoja)
        escribir_hoja(ws, cabeceras, filas, COLORES_HOJA[nombre_hoja])
        print(f"        Hoja '{nombre_hoja}': {len(filas)} filas")

    salida = carpeta_reporte / f"{nombre_reporte}_diccionario.xlsx"
    wb.save(salida)
    print(f"[OK 06] Excel generado → {salida}")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: python 06_generar_excel.py <carpeta_temp> <carpeta_reporte> <nombre_reporte>")
        sys.exit(1)

    generar_excel(
        carpeta_temp    = Path(sys.argv[1]),
        carpeta_reporte = Path(sys.argv[2]),
        nombre_reporte  = sys.argv[3]
    )
