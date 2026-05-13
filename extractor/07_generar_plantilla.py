"""Genera el informe final usando la plantilla institucional.

Toma:
  - extractor/OUTPUT/<reporte>_diccionario.xlsx
  - extractor/OUTPUT/<reporte>_docx_acompanante.xlsx
  - ZDCMTOS/Plantilla Levantamiento de informacion de informes.xlsx

Genera:
  - extractor/OUTPUT/<reporte>_informe.xlsx

Uso:
  python 07_generar_plantilla.py BITACORA
"""

import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR 07] Falta openpyxl. Instalar con: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = ROOT / "ZDCMTOS" / "Plantilla Levantamiento de informacion de informes.xlsx"
OUTPUT_DIR = Path(__file__).resolve().parent / "OUTPUT"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FIELD_LABELS = {
    "organizacion": "ORGANIZACIÓN",
    "nombre_informe": "NOMBRE DE INFORME",
    "area": "AREA",
    "frecuencia_generacion": "FRECUENCIA DE GENERACION",
    "fecha": "FECHA",
    "autor": "AUTOR",
    "objetivo": "OBJETIVO | ¿Qué debe lograr el informe?",
    "dirigido_a": "DIRIGIDO A | ¿A quien va dirigido el informe?",
    "frecuencia_generacion_2": "FRECUENCIA DE GENERACION | ¿Cuál es la frecuencia de generación del informe?",
    "fuentes_utilizadas": "FUENTES UTILIZADAS | ¿Cuáles son las fuentes donde se extrae la informacion?",
    "formato_entrega": "FORMATO DE ENTREGA | ¿Cuál es el formato en que se entrega?",
    "calidad_datos": "CALIDAD DE LOS DATOS | ¿Quien confirma la calidad de los datos del informe?",
}


def leer_acompanante(ruta: Path) -> dict:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    campos = {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or fila[0] is None:
            continue
        clave = str(fila[0]).strip()
        valor = fila[4] if len(fila) >= 5 else None
        if valor is None or str(valor).strip() == "":
            valor = fila[1] if len(fila) >= 2 else None
        if valor is not None:
            campos[clave.lower()] = str(valor).strip()
    return campos


def buscar_fila_por_label(ws, texto: str):
    for fila in ws.iter_rows(min_row=1, max_col=2, values_only=False):
        celda_label = fila[1]
        if celda_label.value and isinstance(celda_label.value, str) and texto.lower() in celda_label.value.lower():
            return celda_label.row
    return None


def obtener_celda_activa(ws, fila: int, columna: int):
    celda = ws.cell(row=fila, column=columna)
    if not isinstance(celda, openpyxl.cell.cell.MergedCell):
        return celda

    for rango in ws.merged_cells.ranges:
        if fila >= rango.min_row and fila <= rango.max_row and columna >= rango.min_col and columna <= rango.max_col:
            return ws.cell(row=rango.min_row, column=rango.min_col)
    return celda


def escribir_valor(ws, label: str, valor, columna: int = 3):
    fila = buscar_fila_por_label(ws, label)
    if fila is None:
        return False
    celda = obtener_celda_activa(ws, fila, columna)
    celda.value = valor
    return True


def crear_hoja_diccionario(wb_dest: openpyxl.Workbook, ruta_diccionario: Path, nombre_reporte: str):
    wb_source = openpyxl.load_workbook(ruta_diccionario, data_only=True)
    hoja_dest = wb_dest.create_sheet(title=f"{nombre_reporte}_diccionario")
    fila_dest = 1

    for hoja_nombre in wb_source.sheetnames:
        ws_source = wb_source[hoja_nombre]
        hoja_dest.cell(row=fila_dest, column=1, value=f"Hoja: {hoja_nombre}")
        hoja_dest.cell(row=fila_dest, column=1).font = openpyxl.styles.Font(bold=True)
        fila_dest += 1

        headers = list(next(ws_source.iter_rows(min_row=1, max_row=1, values_only=True)))
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                hoja_dest.cell(row=fila_dest, column=col_idx, value=header)
                hoja_dest.cell(row=fila_dest, column=col_idx).font = openpyxl.styles.Font(bold=True)
            fila_dest += 1

        for fila in ws_source.iter_rows(min_row=2, values_only=True):
            if all(valor is None for valor in fila):
                continue
            for col_idx, valor in enumerate(fila, start=1):
                hoja_dest.cell(row=fila_dest, column=col_idx, value=valor)
            fila_dest += 1

        fila_dest += 2

    ajustar_ancho_columnas(hoja_dest)


def ajustar_ancho_columnas(ws):
    for columna in ws.columns:
        max_ancho = 0
        for celda in columna:
            if celda.value is not None:
                max_ancho = max(max_ancho, len(str(celda.value)))
        ancho = min(max_ancho + 2, 50)
        ws.column_dimensions[get_column_letter(columna[0].column)].width = ancho


def generar_informe(nombre_reporte: str):
    if not TEMPLATE_PATH.exists():
        print(f"[ERROR 07] No existe la plantilla: {TEMPLATE_PATH}")
        sys.exit(1)

    ruta_docx = OUTPUT_DIR / f"{nombre_reporte}_docx_acompanante.xlsx"
    ruta_dic = OUTPUT_DIR / f"{nombre_reporte}_diccionario.xlsx"

    if not ruta_docx.exists():
        print(f"[ERROR 07] No existe el archivo de acompañante: {ruta_docx}")
        sys.exit(1)
    if not ruta_dic.exists():
        print(f"[ERROR 07] No existe el diccionario generado: {ruta_dic}")
        sys.exit(1)

    datos = leer_acompanante(ruta_docx)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    valores = {
        "organizacion": datos.get("organizacion", "Metrocali"),
        "nombre_informe": nombre_reporte,
        "area": datos.get("area_solicitante", "Evaluacion de la Operacion"),
        "frecuencia_generacion": datos.get("frecuencia_generacion", ""),
        "fecha": datetime.now().date(),
        "autor": datos.get("responsable", datos.get("autor", "")),
        "objetivo": datos.get("objetivo", ""),
        "dirigido_a": datos.get("dirigido_a", ""),
        "frecuencia_generacion_2": datos.get("frecuencia_generacion", datos.get("frecuencia_generacion_2", "")),
        "fuentes_utilizadas": datos.get("fuentes_utilizadas", ""),
        "formato_entrega": datos.get("formato_entrega", ""),
        "calidad_datos": datos.get("calidad_datos", ""),
    }

    for campo, etiqueta in FIELD_LABELS.items():
        escribir_valor(ws, etiqueta, valores.get(campo, ""))

    crear_hoja_diccionario(wb, ruta_dic, nombre_reporte)

    salida = OUTPUT_DIR / f"{nombre_reporte}_informe.xlsx"
    wb.save(salida)
    print(f"[OK 07] Informe generado: {salida}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python 07_generar_plantilla.py <NOMBRE_REPORTE>")
        sys.exit(1)

    generar_informe(sys.argv[1].upper())
